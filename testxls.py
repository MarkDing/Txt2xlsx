#!python
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter
from openpyxl import style
from openpyxl.style import Color, Fill

import re

reg_data = []


class ParseTxt:
    """Extract register data from text file"""
    def __init__(self,filename):
        global f
        offset = 0
        row = 3  # this row number corresponds to excel row number. Row 3 is first register row in excel file.
        f = open(filename,encoding='utf-8')
        # Module name and Description
        t = f.readline().split('\t')[1].split('(')
        tmp = [t[1].split(')')[0],'','','','','','','','','','',t[0].strip(),2]
        reg_data.append(tmp)
        while True:
            line = f.readline()
            if len(line) == 0:
                break;
            if line.startswith("SFR Definition"):
                (offset,row) = self.parse_reg_tbl(line,offset,row)
        f.close()
#        for i in range(len(reg_data)):
#            print(reg_data[i])
        
    def parse_reg_tbl(self,line,offset,row):
        while True:
            # Register definition row
            if line.startswith("SFR Definition"):
                t = line.split(':')
                tmp = ['',t[0].split()[-1], hex(offset),'','','','','','','',t[1].strip(),'',row]
                offset += 1
                row += 1
                reg_data.append(tmp)

            if line.startswith('Type	'):
                bit_rw = line.split('\t')
                del(bit_rw[0])
                bit_rw[-1] = bit_rw[-1].strip()
                for i in range(len(bit_rw)):
                    if bit_rw[i] == 'R/W':
                        bit_rw[i] = 'RW'
                    if bit_rw[i] == '':
                        bit_rw[i] = bit_rw[i-1]
#                print(bit_rw)
            if line.startswith('Reset	'):
                bit_rst = line.split('\t')
                del(bit_rst[0])
                bit_rst[-1] = bit_rst[-1].strip()
#                print(bit_rst)
                
#                        tmp = ['','','','','','','','','','','','',row]
            # register description    
            if line.startswith('Bit	Name	Function'):
                idx = 0 # bit index
                description_row = row
                while True:
                    line = f.readline()
                    t = line.split('\t')
                    if (len(t[0]) == 1) and (t[0].isnumeric()) :# 7->
                        tmp = ['','','',t[0],t[1],bit_rw[idx],bit_rst[idx],'','','',t[2].strip(),'',row]
                        description_row = row
                        idx += 1
                        row += 1
                        reg_data.append(tmp)
#                        print(tmp)
                    elif(len(t[0]) == 3) and t[0][0].isnumeric() and (t[0][1] == ':') and t[0][2].isnumeric():# 7:0->
                        bit_num = int(t[0][0])
                        bit_name = t[1].split('[')[0]
                        short_name = t[2].strip()
                        description_row = row
                        while True:
                            tmp = ['','','',bit_num,bit_name,bit_rw[idx],bit_rst[idx],'','','',short_name,'',row]
                            short_name = ''
                            idx += 1
                            row += 1
                            reg_data.append(tmp)
                            bit_num -= 1
                            if bit_num < int(t[0][2]):
                               break
                    else:
                        t = line.split(':')
                        if t[0][0].isnumeric() and ((t[0][-1] == 'x') or (t[0][-1].isnumeric())): # 1x:
                            if (t[0][-1] == 'x'):
                                tmp = ['','','','','','','','???',t[0],'','',t[1].strip(),row]
                            else:    
                                tmp = ['','','','','','','','???',int(t[0],2),'','',t[1].strip(),row]
                            row +=1
                            reg_data.append(tmp)
                        else: # description
                            reg_data[description_row - 2][11] += line
                        
                    # two CR means reach end of this register definition. 
                    if line.startswith('\n'):
                        line = f.readline()
                        if line.startswith('\n'):
#                            print("endof register")
                            return (offset,row)
            line = f.readline()
            if len(line) == 0:
                break;
        

class WriteXlsx:
    """Write regs data into xlsx file"""
    def __init__(self,dest_filename,reg):
#        print("WriteXlsx class init");
        wb = Workbook()
        global ws
        ws = wb.worksheets[0]
#        ws.title = 'version'
#        ws = wb.create_sheet()
        ws.title = "Base_alias"
        ws._freeze_panes = 'A2'
        self.write_regs(reg)
        wb.save(filename = dest_filename)

    def write_row(self,row_data):
        row = row_data[-1]
        high_light = 0
        if row_data[1] != '':  # register name field
            high_light = 1
        for i in range(0,len(row_data)-1):
            col = get_column_letter(i+1)
            if high_light == 1:
                ws.cell('%s%s'%(col, row)).style.fill.fill_type = Fill.FILL_SOLID
                ws.cell('%s%s'%(col, row)).style.fill.start_color.index = Color.YELLOW
            ws.cell('%s%s'%(col, row)).value = row_data[i]
            
    def write_regs(self,reg):
        row1_data = [
            'Module','Register','Offset','Bit Num','Bit','Bit Access','Bit Field Reset',
            'Enum Name','Enum Value','Special','Short Name','Description',
            ]
        # write first line
        for col_idx in range(0, len(row1_data)):
            col = get_column_letter(col_idx+1)
            ws.cell('%s%s'%(col, 1)).value = row1_data[col_idx]
            ws.cell('%s%s'%(col, 1)).style.font.bold = True
        for i in range(0,len(reg)):
            self.write_row(reg[i])

a = '110101'
print(int(a,2))

print("Start convertion")
src_filename = r'SPI.txt'
pt = ParseTxt(src_filename)

dest_filename = r'empty_book.xlsx'
wx = WriteXlsx(dest_filename,reg_data)
del(reg_data[:])
print("Excel file generated")
