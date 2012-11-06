#!python
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter
from openpyxl import style
from openpyxl.style import Color, Fill, Borders
import sys

reg_data = []


class ParseTxt:
    """Extract register data from text file"""
    def __init__(self,filename):
        global f
        offset = 0
        row = 3  # this row number corresponds to excel row number. Row 3 is first register row in excel file.
        f = open(filename,encoding='utf-8')
        # Module name and Description
        tmp = [filename.split('.')[0],'','','','','','','','','','','???',2]
        reg_data.append(tmp)
        while True:
            line = f.readline()
            if len(line) == 0:
                break;
            if (line.startswith("SFR Definition") or line.startswith("Internal Register Definition")):
                (offset,row) = self.parse_reg_tbl(line,offset,row)
        f.close()
#        for i in range(len(reg_data)):
#            print(reg_data[i])
        
    def parse_reg_tbl(self,line,offset,row):
        while True:
            # Register definition row. EX: SFR Definition 24.2. SPI0CN: SPI0 Control
            if (line.startswith("SFR Definition") or line.startswith("Internal Register Definition")):
                t = line.split(':')
                print(t)
                tmp = ['',t[0].split()[-1], hex(offset),'','','','','','','',t[1].strip(),'',row]
                offset += 1
                row += 1
                reg_data.append(tmp)
            # Bit field type. EX: Type	R/W	R/W	R/W	R/W	R/W		R	R/W
            if line.startswith('Type	'):
                bit_rw = line.split('\t')
                del(bit_rw[0])
                bit_rw[-1] = bit_rw[-1].strip()
                for i in range(len(bit_rw)):
                    if bit_rw[i] == 'R/W':
                        bit_rw[i] = 'RW'
                    if bit_rw[i] == '':
                        bit_rw[i] = bit_rw[i-1]
            # Bit field value. EX: Reset 0	0	0	0	0	1	1	0
            if line.startswith('Reset	'):
                bit_rst = line.split('\t')
                del(bit_rst[0])
                bit_rst[-1] = bit_rst[-1].strip()
                
            # register description
            # EX: Bit	Name	Function		
            if line.startswith('Bit	Name	Function') or line.startswith('Bit	Name	Description'):
                idx = 0 # bit index
                start_row = row
                description_row = row
                while True:
                    line = f.readline()
                    t = line.split('\t')
#                    print(t)
                    # EX: 7	SPIF	SPI0 Interrupt Flag.
                    if (len(t[0]) == 1) and (t[0].isnumeric()) :# 7->
                        tmp = ['','','',t[0],t[1],bit_rw[idx],bit_rst[idx],'','','',t[2].strip(),'',row]
                        description_row = row
                        idx += 1
                        row += 1
                        reg_data.append(tmp)
#                        print(tmp)
                    # EX: 3:2	NSSMD[1:0]	Slave Select Mode.
                    elif(len(t[0]) == 3) and t[0][0].isnumeric() and (t[0][1] == ':') and t[0][2].isnumeric():# 7:0->
                        bit_num = int(t[0][0])
                        bit_name = t[1].split('[')[0]
                        short_name = t[2].strip()
                        description_row = row
                        while True:
                            tmp = ['','','',bit_num,bit_name,bit_rw[idx],bit_rst[idx],'','','','','',row]
#                            print(tmp)
                            idx += 1
                            row += 1
                            reg_data.append(tmp)
                            bit_num -= 1
                            if bit_num < int(t[0][2]):
                                description_row = row - 1
#                                print(short_name)
                                reg_data[description_row - 2][10] = short_name
                                break
                    else:
                        t = line.split(':')
#                        print(t)
                        if t[0][0].isnumeric() and ((t[0][-1] == 'x') or (t[0][-1].isnumeric())): # 1x:
                            # EX: 1x: 4-Wire Single-Master Mode. NSS signal is mapped as an output from the device and will assume the value of NSSMD0.		
                            if (t[0][-1] == 'x'):
                                # EX: 1xx: ADC0 conversion initiated on rising edge of CNVSTR.
                                if (t[0][-2] == 'x'):
                                    tmp1x = t[0][0:-2] + '00'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                                    tmp1x = t[0][0:-2] + '01'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                                    tmp1x = t[0][0:-2] + '10'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                                    tmp1x = t[0][0:-2] + '11'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                                else:    
                                    tmp1x = t[0][0:-1] + '0'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                                    tmp1x = t[0][0:-1] + '1'
                                    tmp = ['','','','','','','','???',int(tmp1x,2),'','',t[1].strip('\n'),row]
                                    row +=1
                                    reg_data.append(tmp)
                            # EX: 00: 3-Wire Slave or 3-Wire Master Mode. NSS signal is not routed to a port pin.                                
                            else:    
                                tmp = ['','','','','','','','???',int(t[0],2),'','',t[1].strip('\n'),row]
                                row +=1
                                reg_data.append(tmp)
                        else: # description. EX: Selects between the following NSS operation modes:
                            # reg_data start from row 2 in excel file, so we minus 2 here.
                            reg_data[description_row - 2][11] += line.strip('\n')
                        
                    # two CR means reach end of this register definition. 
                    if line.startswith('\n'):
                        line = f.readline()
                        if line.startswith('\n'):
                            num_rows = row - start_row
                            row_tmp = [x for x in range(0,num_rows)]
                            tmp_idx = num_rows
                            reg_idx = start_row - 2
#                            print('row=%s,start_row=%s'%(row,start_row))
                            while True:
                                cnt = 1
                                idx = reg_idx + 1
                                while True:
                                    if idx == (row - 2):
                                        break
                                    if reg_data[idx][3] == '':
                                        idx += 1
                                        cnt += 1
                                    else:
                                        break
                                tmp_idx -= cnt
#                                print('********')
#                                print('tmp_idx=%s,cnt=%s,reg_idx=%s'%(tmp_idx,cnt,reg_idx))
                                for i in range(0,cnt):
                                    row_tmp[tmp_idx + i] = reg_data[reg_idx+i][12]
                                reg_idx += cnt
                                if reg_idx == (row - 2):
                                    break
                            for i in range(0, num_rows):
                                old_idx = row_tmp[i] - 2
                                reg_data[old_idx][12] = i + start_row
#                                print('old_idx=%d  i+start_row = %d'%(old_idx,i+start_row))
                            return (offset,row)
            line = f.readline()
            if len(line) == 0:
                break;
        

class WriteXlsx:
    """Write regs data into xlsx file"""
    def __init__(self,dest_filename,reg):
#        print("WriteXlsx class init");
        wb = Workbook()
        global ws, ws0
        ws0 = wb.worksheets[0]
        ws0.title = 'Version'
        self.fill_sheet0(dest_filename.split('.')[0])
        ws = wb.create_sheet()
        ws.title = "Base_Alias"
        ws._freeze_panes = 'A2'
        self.write_regs(reg)
        wb.save(filename = dest_filename)

    def fill_sheet0(self,name):
#        ws0.cell('A2').style.borders.bottom = Borders.DIAGONAL_DOWN
        ws0.cell('A3').value = 'Variable section'
        ws0.cell('A4').value = 'AliasParserVersion'
        ws0.cell('B4').value = '2'
        ws0.cell('A5').value = 'CoreType'
        ws0.cell('B5').value = 'CIP51'
        ws0.cell('A6').value = 'End Variable section'
        ws0.cell('A9').value = 'Change List - Newest on Top'
        ws0.cell('A10').value = '11/05/2012'
        ws0.cell('B10').value = 'MD'
        ws0.cell('C10').value = 'Initial Kylin ' + name
        
    def write_row(self,row_data):
        row = row_data[-1]
        high_light = 0
        if row_data[1] != '':  # register name field
            high_light = 1
        for i in range(0,len(row_data)-1):
            col = get_column_letter(i+1)
            if high_light == 1:
                ws.cell('%s%s'%(col, row)).style.fill.fill_type = Fill.FILL_SOLID
                ws.cell('%s%s'%(col, row)).style.fill.start_color.index = Color.YELLOW
            ws.cell('%s%s'%(col, row)).value = row_data[i]
            
    def write_regs(self,reg):
        row1_data = [
            'Module','Register','Offset','Bit Num','Bit','Bit Access','Bit Field Reset',
            'Enum Name','Enum Value','Special','Short Name','Description',
            ]
        # write first line
        for col_idx in range(0, len(row1_data)):
            col = get_column_letter(col_idx+1)
            ws.cell('%s%s'%(col, 1)).value = row1_data[col_idx]
            ws.cell('%s%s'%(col, 1)).style.font.bold = True
        for i in range(0,len(reg)):
            self.write_row(reg[i])

a = '110101'
#print(int(a,2))


args = sys.argv[1:]
if args == []:
    print('No filename')
else:
    src_filename = args[0].upper()
    print("Start convertion")
    pt = ParseTxt(src_filename)

    dest_filename = src_filename.split('.')[0]+'.xlsx'
    wx = WriteXlsx(dest_filename,reg_data)
    del(reg_data[:])
    print("Excel file generated")
